import time
import os
import logging
import traceback
import pickle
from pathlib import Path
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta

class LinkedinLogin:
    def __init__(self, browser):
        # Load environment variables
        env_file = os.path.join(Path(__file__).parent, ".env")
        load_dotenv(env_file)
        self.browser = browser
        self.username = os.getenv("LINKEDIN_USERNAME")
        self.password = os.getenv("LINKEDIN_PASSWORD")
        self.cookies_file = "linkedin_cookies.pkl"
        self.comments_file = "comments.txt"
        self.comments = self._load_comments()
        self.historical_file = "historical_posts.xlsx"
        self.processed_urls, self.used_comments = self._load_processed_urls_and_comments()
        self.new_credentials = os.getenv("NEW_CREDENTIALS", "false").lower() == "true"
        logging.info(f'Username: {self.username}')
        self.last_break_time = datetime.now()

    def _load_comments(self):
        # Load comments from the comments.txt file
        try:
            with open(self.comments_file, "r") as file:
                comments = file.readlines()
            return [comment.strip() for comment in comments]
        except Exception as e:
            logging.error(f"Error loading comments: {e}")
            logging.error(traceback.format_exc())
            return []

    def _load_processed_urls_and_comments(self):
        # Load processed URLs and used comments from the historical_posts.xlsx file
        try:
            processed_urls = set()
            used_comments = set()
            if os.path.exists(self.historical_file):
                wb = load_workbook(self.historical_file)
                sheet = wb.active if wb.sheetnames else None
                if sheet:
                    for row in sheet.iter_rows(values_only=True):
                        processed_urls.add(row[0])
                        used_comments.add(row[1])
            return processed_urls, used_comments
        except Exception as e:
            logging.error(f"Error loading processed URLs and comments: {e}")
            logging.error(traceback.format_exc())
            return set(), set()

    def save_processed_url(self, url, comment):
        # Save the processed URL and comment to the historical_posts.xlsx file
        try:
            wb = load_workbook(self.historical_file) if os.path.exists(self.historical_file) else Workbook()
            if not wb.sheetnames:
                sheet = wb.create_sheet(title="Processed Notifications")
                sheet.append(["Notification URL", "Comment", "Timestamp"])
            else:
                sheet = wb.active
            sheet.append([url, comment, time.strftime("%Y-%m-%d %H:%M:%S")])
            wb.save(self.historical_file)
            logging.info(f"Logged processed notification: {url}")
        except Exception as e:
            logging.error(f"Error logging processed notification: {e}")
            logging.error(traceback.format_exc())

    def load_cookies(self):
        # Load cookies from the cookies.pkl file
        try:
            if os.path.exists(self.cookies_file) and not self.new_credentials:
                with open(self.cookies_file, "rb") as file:
                    cookies = pickle.load(file)
                    for cookie in cookies:
                        self.browser.add_cookie(cookie)
                    logging.info("Cookies loaded successfully.")
                    return True
            return False
        except Exception as e:
            logging.error(f"Error loading cookies: {e}")
            logging.error(traceback.format_exc())
            return False

    def save_cookies(self):
        # Save cookies to the cookies.pkl file
        try:
            with open(self.cookies_file, "wb") as file:
                pickle.dump(self.browser.get_cookies(), file)
            logging.info("Cookies saved successfully.")
        except Exception as e:
            logging.error(f"Error saving cookies: {e}")
            logging.error(traceback.format_exc())

    def check_break_time(self):
        # Check if the script needs to take a break
        current_time = datetime.now()
        elapsed_time = current_time - self.last_break_time
        if elapsed_time > timedelta(hours=1):
            logging.info("Taking a 10-minute break...")
            time.sleep(600)  # 10-minute break
            self.last_break_time = datetime.now()

    def login_and_open_notifications(self):
        # Log in to LinkedIn and open the notifications tab
        try:
            self.browser.get("https://www.linkedin.com")
            if not self.load_cookies() or self.new_credentials:
                self.browser.get("https://www.linkedin.com/login")
                self._perform_login()
                self.save_cookies()
            else:
                self.browser.get("https://www.linkedin.com")
                time.sleep(2)
                self.browser.refresh()
                if not self._is_logged_in():
                    self.browser.get("https://www.linkedin.com/login")
                    self._perform_login()
                    self.save_cookies()

            # Navigate to notifications
            notifications_tab = WebDriverWait(self.browser, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='Notifications']"))
            )
            self.browser.execute_script("arguments[0].click();", notifications_tab)
            time.sleep(2)

            logging.info("Successfully navigated to Notifications tab!")

            # Process all "New from" newsletter notifications
            notification_index = 1
            while True:
                self.check_break_time()  # Check if it's time to take a break
                try:
                    notification_xpath = f"(//a[contains(@class, 'nt-card__headline')])[{notification_index}]"
                    notification = WebDriverWait(self.browser, 20).until(
                        EC.presence_of_element_located((By.XPATH, notification_xpath))
                    )

                    notification_url = notification.get_attribute("href")
                    if self.is_newsletter_notification(notification) and notification_url not in self.processed_urls:
                        logging.info(f"Newsletter notification URL: {notification_url}")

                        # Open the URL in a new tab
                        self.browser.execute_script("window.open(arguments[0], '_blank');", notification_url)
                        logging.info(f"Successfully opened newsletter notification {notification_index} in a new tab!")
                        time.sleep(5)

                        # Switch to the new tab
                        self.browser.switch_to.window(self.browser.window_handles[1])

                        # Perform the actions (like, comment, close tab)
                        self.like_post()
                        self.comment_on_post(notification_url)

                        # Wait for 5 seconds before closing the current tab
                        time.sleep(5)

                        # Close the current tab and switch back to the first tab
                        self.browser.close()
                        self.browser.switch_to.window(self.browser.window_handles[0])
                        logging.info(f"Successfully processed newsletter notification {notification_index} and switched back to the notifications tab!")

                    # Wait for 2 seconds before processing the next notification
                    time.sleep(2)

                    notification_index += 1

                    # Scroll down to load more notifications if necessary
                    if notification_index % 5 == 0:
                        self.scroll_to_bottom()

                except Exception as e:
                    logging.error(f"Error processing notification {notification_index}: {e}")
                    logging.error(traceback.format_exc())
                    # Close the tab and continue with the next notification
                    if len(self.browser.window_handles) > 1:
                        self.browser.close()
                        self.browser.switch_to.window(self.browser.window_handles[0])
                    notification_index += 1
                    continue

                # Break the loop if no new notifications to process
                if notification_index > len(self.browser.find_elements(By.XPATH, "//a[contains(@class, 'nt-card__headline')]")):
                    logging.info("No new notifications remaining to open.")
                    break

        except Exception as e:
            logging.error(f"Error during login and navigation: {e}")
            logging.error(traceback.format_exc())
            raise e

    def _perform_login(self):
        # Perform login using username and password
        try:
            self.browser.implicitly_wait(10)
            time.sleep(2)
            username_field = self.browser.find_element(By.ID, "username")
            username_field.send_keys(self.username)
            time.sleep(2)

            password_field = self.browser.find_element(By.ID, "password")
            password_field.send_keys(self.password)
            time.sleep(2)
            password_field.send_keys(Keys.RETURN)

            # Wait for 2FA input (manually enter OTP if needed)
            logging.info("Please complete the login process, including any 2FA if required...")
            WebDriverWait(self.browser, 60).until(
                EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Search']"))
            )
            logging.info("Login completed successfully.")
            time.sleep(2)
        except Exception as e:
            logging.error(f"Error during login: {e}")
            logging.error(traceback.format_exc())
            raise e

    def _is_logged_in(self):
        # Check if the user is logged in
        try:
            self.browser.get("https://www.linkedin.com/feed/")
            WebDriverWait(self.browser, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Search']"))
            )
            return True
        except Exception:
            return False

    def is_newsletter_notification(self, notification):
        # Check if the notification is a newsletter
        try:
            notification_text = notification.text
            return notification_text.startswith("New from")
        except Exception as e:
            logging.error(f"Error checking if notification is a newsletter: {e}")
            logging.error(traceback.format_exc())
            return False

    def like_post(self):
        # Like the post
        try:
            like_button = WebDriverWait(self.browser, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'React Like')]"))
            )
            self.browser.execute_script("arguments[0].click();", like_button)
            logging.info("Successfully liked the post!")
            time.sleep(5)
        except Exception as e:
            logging.error(f"Error during liking the post: {e}")
            logging.error(traceback.format_exc())

    def comment_on_post(self, notification_url):
        # Comment on the post
        try:
            comment_button = WebDriverWait(self.browser, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Comment')]"))
            )
            self.browser.execute_script("arguments[0].click();", comment_button)
            logging.info("Successfully opened the comment input!")
            time.sleep(5)

            comment_input = WebDriverWait(self.browser, 20).until(
                EC.presence_of_element_located((By.XPATH, "//div[@role='textbox']"))
            )
            if self.comments:
                comment = self.comments.pop(0)
                while comment in self.used_comments:
                    if not self.comments:
                        logging.warning("No more unique comments left to post.")
                        return
                    comment = self.comments.pop(0)
                comment_input.send_keys(comment)
                logging.info(f"Successfully entered the comment: {comment}")

                post_button = WebDriverWait(self.browser, 20).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@class='comments-comment-box__submit-button mt3 artdeco-button artdeco-button--1 artdeco-button--primary ember-view']"))
                )
                self.browser.execute_script("arguments[0].click();", post_button)
                logging.info("Successfully submitted the comment!")
                time.sleep(5)
                self.save_processed_url(notification_url, comment)
                self.used_comments.add(comment)
            else:
                logging.warning("No more comments left to post.")
        except Exception as e:
            logging.error(f"Error during commenting on the post: {e}")
            logging.error(traceback.format_exc())

    def scroll_to_bottom(self):
        # Scroll to the bottom of the page to load more notifications
        try:
            self.browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            logging.info("Scrolled to the bottom of the page.")
        except Exception as e:
            logging.error(f"Error while scrolling to the bottom: {e}")
            logging.error(traceback.format_exc())

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    # Configure Chrome options
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-notifications")

    # Set up Chrome driver
    try:
        driver_path = ChromeDriverManager().install()
        browser = webdriver.Chrome(service=Service(driver_path), options=chrome_options)

        # Initialize LinkedIn login and interaction
        linkedin_login = LinkedinLogin(browser)
        linkedin_login.login_and_open_notifications()

    except Exception as e:
        logging.error(f"Main script error: {e}")
        logging.error(traceback.format_exc())
    finally:
        if browser:
            browser.quit()
